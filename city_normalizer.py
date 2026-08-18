"""Нормализация столбца «город» к коротким названиям районов Краснодарского края."""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

from krasnodar_districts import (
    LOCALITY_TO_DISTRICTS,
    OFFICIAL_DISTRICTS,
    SHORT_DISTRICT_LIST,
    to_official,
    to_short,
)

DEFAULT_THRESHOLD = 82
FUZZY_CERTAIN_SCORE = 95
SUPPORTED_EXCEL = {".xlsx", ".xls", ".xlsm"}
SUPPORTED_CSV = {".csv"}
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1251", "windows-1251")
USER_DICT_PATH = Path(__file__).resolve().parent / "user_localities.json"

STATUS_CERTAIN = "уверен"
STATUS_UNCERTAIN = "не уверен"
STATUS_UNMATCHED = "не сопоставлено"
STATUS_EMPTY = "пусто"
STATUS_COL = "уверенность"
VARIANTS_COL = "варианты"

COLUMN_ALIASES = (
    "населённый пункт",
    "населенный пункт",
    "город",
    "city",
    "адрес",
    "address",
    "местность",
    "район",
    "нп",
)


def _fold(text: str) -> str:
    return " ".join(text.lower().replace("ё", "е").split())


_OFFICIAL_FOLD = {_fold(name): name for name in OFFICIAL_DISTRICTS}
_SHORT_FOLD = {_fold(name): name for name in SHORT_DISTRICT_LIST}

_PREFIX_RE = re.compile(
    r"""^
    (?:
        муниципальный\s+округ\s+город(?:-курорт|-герой)?
        |городской\s+округ\s+город(?:-курорт|-герой)?
        |муниципальный\s+округ
        |городской\s+округ
        |муниципальный\s+район
        |город-курорт
        |город-герой
        |г\.-к\.?
        |ст\.-ца\.?
        |р\.\s*п\.
        |к/п
        |г/о
        |м\.р\.
        |пгт\.?
        |пос[её]лок
        |станица
        |деревня
        |хутор
        |село
        |аул
        |пос\.
        |рп\.?
        |кп\.?
        |город
        |г\.
        |п\.
        |с\.
        |ст\.
        |х\.
        |а\.
        |д\.
    )
    [\s.:\-]*
    """,
    re.IGNORECASE | re.VERBOSE,
)


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def strip_prefix(value: str) -> str:
    """Убирает административные префиксы (г., ст., пгт, городской округ …)."""
    text = str(value).strip().strip("«»\"'`")
    if not text:
        return ""
    previous = None
    while previous != text:
        previous = text
        text = _PREFIX_RE.sub("", text, count=1).strip(" .,;-")
    return " ".join(text.split())


def load_user_localities() -> dict[str, str]:
    """Пользовательский справочник: свёрнутое имя НП → короткое имя района."""
    if not USER_DICT_PATH.exists():
        return {}
    try:
        raw = json.loads(USER_DICT_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    result: dict[str, str] = {}
    for key, value in raw.items():
        folded = _fold(str(key))
        short = to_short(str(value))
        if folded and short in SHORT_DISTRICT_LIST:
            result[folded] = short
    return result


def save_user_localities(mapping: dict[str, str]) -> None:
    clean = {
        _fold(key): to_short(value)
        for key, value in mapping.items()
        if _fold(key) and to_short(value) in SHORT_DISTRICT_LIST
    }
    USER_DICT_PATH.write_text(
        json.dumps(clean, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def upsert_user_locality(name: str, district: str) -> str:
    """Добавляет или обновляет НП в пользовательском справочнике."""
    cleaned = strip_prefix(name) or str(name).strip()
    folded = _fold(cleaned)
    if not folded:
        raise ValueError("Пустое название населённого пункта.")
    short = to_short(district)
    if short not in SHORT_DISTRICT_LIST:
        raise ValueError(f"Неизвестный район: {district}")
    mapping = load_user_localities()
    mapping[folded] = short
    save_user_localities(mapping)
    return cleaned


def delete_user_locality(name: str) -> None:
    mapping = load_user_localities()
    mapping.pop(_fold(strip_prefix(name) or name), None)
    save_user_localities(mapping)


def merged_locality_map(user_map: dict[str, str] | None = None) -> dict[str, list[str]]:
    """Встроенный справочник + пользовательские записи (они снимают неоднозначность)."""
    merged: dict[str, list[str]] = {
        key: list(districts) for key, districts in LOCALITY_TO_DISTRICTS.items()
    }
    for key, short in (user_map if user_map is not None else load_user_localities()).items():
        official = to_official(short)
        if official:
            merged[key] = [official]
    return merged


def _shorts(officials: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for official in officials:
        short = to_short(official)
        if short not in seen:
            seen.add(short)
            result.append(short)
    return result


@dataclass
class MatchResult:
    value: object
    status: str
    candidates: list[str] = field(default_factory=list)
    cleaned: str = ""
    method: str = "none"
    score: float | None = None

    @property
    def variants_text(self) -> str:
        if self.status != STATUS_UNCERTAIN:
            return ""
        return " / ".join(self.candidates)


def _from_districts(
    officials: list[str],
    *,
    cleaned: str,
    method: str,
    score: float | None = None,
) -> MatchResult:
    shorts = _shorts(officials)
    if not shorts:
        return MatchResult(cleaned, STATUS_UNMATCHED, [], cleaned, "none")
    if len(shorts) == 1:
        certain = method in {"exact", "user", "official"} or (
            method == "fuzzy" and score is not None and score >= FUZZY_CERTAIN_SCORE
        )
        return MatchResult(
            shorts[0],
            STATUS_CERTAIN if certain else STATUS_UNCERTAIN,
            shorts,
            cleaned,
            method,
            score,
        )
    return MatchResult(
        " / ".join(shorts),
        STATUS_UNCERTAIN,
        shorts,
        cleaned,
        method,
        score,
    )


def fuzzy_lookup(
    name: str,
    threshold: int = DEFAULT_THRESHOLD,
    user_map: dict[str, str] | None = None,
) -> MatchResult | None:
    """Нечёткий поиск среди ключей справочника."""
    query = _fold(name)
    if not query:
        return None
    lookup = merged_locality_map(user_map)
    hit = process.extractOne(
        query,
        list(lookup.keys()),
        scorer=fuzz.WRatio,
        score_cutoff=threshold,
    )
    if hit is None:
        return None
    key, score = hit[0], float(hit[1])
    return _from_districts(lookup[key], cleaned=name, method="fuzzy", score=score)


def match_city(
    value: object,
    threshold: int = DEFAULT_THRESHOLD,
    user_map: dict[str, str] | None = None,
) -> MatchResult:
    """Сопоставление значения с коротким названием района и статусом уверенности."""
    if _is_empty(value):
        return MatchResult(value, STATUS_EMPTY, [], "", "empty")

    original = str(value).strip()
    if original.lower() in {"nan", "none", "null", "-"}:
        return MatchResult(original, STATUS_EMPTY, [], original, "empty")

    users = user_map if user_map is not None else load_user_localities()
    cleaned = strip_prefix(original)
    folded_raw = _fold(original)
    folded = _fold(cleaned)

    for key in (folded_raw, folded):
        if key in users:
            short = users[key]
            return MatchResult(short, STATUS_CERTAIN, [short], cleaned, "user")

    for key in (folded_raw, folded):
        if key in _SHORT_FOLD:
            short = _SHORT_FOLD[key]
            return MatchResult(short, STATUS_CERTAIN, [short], cleaned, "official")
        if key in _OFFICIAL_FOLD:
            short = to_short(_OFFICIAL_FOLD[key])
            return MatchResult(short, STATUS_CERTAIN, [short], cleaned, "official")

    lookup = merged_locality_map(users)
    for key in (folded_raw, folded):
        if key in lookup:
            return _from_districts(lookup[key], cleaned=cleaned, method="exact")

    fuzzy = fuzzy_lookup(folded or folded_raw, threshold=threshold, user_map=users)
    if fuzzy is not None:
        fuzzy.cleaned = cleaned
        return fuzzy

    return MatchResult(cleaned or original, STATUS_UNMATCHED, [], cleaned, "none")


def normalize_city(value: object, threshold: int = DEFAULT_THRESHOLD) -> object:
    """Короткое имя района либо очищенное исходное значение."""
    return match_city(value, threshold=threshold).value


def detect_city_column(columns: Iterable[object]) -> str | None:
    """Автоопределение столбца с городом/населённым пунктом."""
    names = [str(col) for col in columns]
    folded = [_fold(name) for name in names]

    for alias in COLUMN_ALIASES:
        alias_fold = _fold(alias)
        for original, current in zip(names, folded):
            if current == alias_fold:
                return original

    for alias in COLUMN_ALIASES:
        alias_fold = _fold(alias)
        for original, current in zip(names, folded):
            if alias_fold in current.split() or alias_fold == current:
                return original
            if alias_fold in current and alias_fold != "нп":
                return original
    return names[0] if names else None


def _suffix(path: str | Path) -> str:
    return Path(path).suffix.lower()


def read_file(
    source: str | Path | BinaryIO,
    filename: str | None = None,
) -> pd.DataFrame:
    """Читает xlsx / xls / xlsm / csv (utf-8, utf-8-sig, windows-1251)."""
    name = filename
    if name is None and isinstance(source, (str, Path)):
        name = str(source)
    if not name:
        raise ValueError("Не удалось определить имя файла.")

    ext = _suffix(name)
    if ext in SUPPORTED_EXCEL:
        return _read_excel(source, ext)
    if ext in SUPPORTED_CSV:
        return _read_csv(source)
    raise ValueError(
        f"Неподдерживаемый формат «{ext}». Допустимы: xlsx, xls, xlsm, csv."
    )


def _as_buffer(source: str | Path | BinaryIO) -> tuple[bool, object]:
    if isinstance(source, (str, Path)):
        return True, source
    if hasattr(source, "seek"):
        source.seek(0)
    return False, source


def _read_excel(source: str | Path | BinaryIO, ext: str) -> pd.DataFrame:
    engine = "xlrd" if ext == ".xls" else "openpyxl"
    _, handle = _as_buffer(source)
    return pd.read_excel(handle, engine=engine)


def _read_csv(source: str | Path | BinaryIO) -> pd.DataFrame:
    is_path, handle = _as_buffer(source)
    raw: bytes
    if is_path:
        raw = Path(handle).read_bytes()
    else:
        raw = handle.read()
        if isinstance(raw, str):
            raw = raw.encode("utf-8")

    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        for sep in (",", ";", "\t"):
            try:
                df = pd.read_csv(
                    io.BytesIO(raw),
                    encoding=encoding,
                    sep=sep,
                    dtype=str,
                    keep_default_na=False,
                    na_values=["", "NA", "NaN", "nan", "None"],
                )
            except (UnicodeDecodeError, pd.errors.ParserError) as exc:
                last_error = exc
                continue
            if df.shape[1] == 1 and sep == "," and ";" in raw.decode(
                encoding, errors="ignore"
            ):
                continue
            return df
    raise ValueError(
        "Не удалось прочитать CSV. Проверьте кодировку и разделитель."
    ) from last_error


FILL_UNCERTAIN = PatternFill("solid", fgColor="FFF3CD")
FILL_UNMATCHED = PatternFill("solid", fgColor="F8D7DA")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_BOLD = Font(bold=True)


def _style_worksheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    status_idx = None
    city_idx = None
    if STATUS_COL in df.columns:
        status_idx = list(df.columns).index(STATUS_COL) + 1
    for name in df.columns:
        folded = _fold(str(name))
        if folded in {"город", "city", "населенный пункт", "населённый пункт"}:
            city_idx = list(df.columns).index(name) + 1
            break
    if city_idx is None and len(df.columns) > 0:
        city_idx = 1

    statuses = df[STATUS_COL] if STATUS_COL in df.columns else []
    for row_idx, status in enumerate(statuses, start=2):
        fill = None
        if status == STATUS_UNCERTAIN:
            fill = FILL_UNCERTAIN
        elif status == STATUS_UNMATCHED:
            fill = FILL_UNMATCHED
        if fill is None:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
        if city_idx:
            ws.cell(row=row_idx, column=city_idx).font = FONT_BOLD
        if status_idx:
            ws.cell(row=row_idx, column=status_idx).font = FONT_BOLD

    for col_idx, name in enumerate(df.columns, start=1):
        series = df[name].astype(str)
        samples = [len(str(name))] + [len(v) for v in series.head(80)]
        width = min(max(samples, default=10) + 2, 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_file(df: pd.DataFrame, path: str | Path) -> Path:
    """Сохраняет таблицу в xlsx / xlsm / csv. Для .xls пишет .xlsx."""
    out = Path(path)
    ext = out.suffix.lower()
    out.parent.mkdir(parents=True, exist_ok=True)

    if ext == ".csv":
        df.to_csv(out, index=False, encoding="utf-8-sig")
        return out
    if ext in {".xlsx", ".xlsm", ".xls"}:
        if ext == ".xls":
            out = out.with_suffix(".xlsx")
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Нормализация")
            _style_worksheet(writer.sheets["Нормализация"], df)
        return out
    raise ValueError(
        f"Неподдерживаемый формат записи «{ext}». Допустимы: xlsx, xlsm, csv."
    )


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Нормализация")
        _style_worksheet(writer.sheets["Нормализация"], df)
    return buffer.getvalue()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def style_preview(df: pd.DataFrame):
    """Pandas Styler: жёлтый — не уверен, красный — не сопоставлено."""

    def _row_style(row: pd.Series) -> list[str]:
        status = row.get(STATUS_COL, "")
        if status == STATUS_UNCERTAIN:
            color = "background-color: #fff3cd; font-weight: bold"
        elif status == STATUS_UNMATCHED:
            color = "background-color: #f8d7da"
        else:
            return [""] * len(row)
        return [color] * len(row)

    return df.style.apply(_row_style, axis=1)


def _sort_by_confidence(df: pd.DataFrame) -> pd.DataFrame:
    order = {
        STATUS_UNCERTAIN: 0,
        STATUS_UNMATCHED: 1,
        STATUS_CERTAIN: 2,
        STATUS_EMPTY: 3,
    }
    if STATUS_COL not in df.columns:
        return df
    ranked = df[STATUS_COL].map(lambda x: order.get(x, 9))
    return df.assign(_rank=ranked).sort_values("_rank", kind="mergesort").drop(columns="_rank")


@dataclass
class ProcessResult:
    df: pd.DataFrame
    column: str
    total: int
    changed: int
    unmatched: int
    uncertain: int
    certain: int
    exact: int
    fuzzy: int
    unmatched_rows: pd.DataFrame = field(default_factory=pd.DataFrame)
    uncertain_rows: pd.DataFrame = field(default_factory=pd.DataFrame)
    output_path: Path | None = None
    matches: list[MatchResult] = field(default_factory=list)
    originals: list[object] = field(default_factory=list)


def process_file(
    source: str | Path | BinaryIO | None = None,
    *,
    filename: str | None = None,
    output: str | Path | None = None,
    column: str | None = None,
    add_original: bool = False,
    dry_run: bool = False,
    threshold: int = DEFAULT_THRESHOLD,
    df: pd.DataFrame | None = None,
) -> ProcessResult:
    """Полный цикл: чтение → нормализация столбца → статистика → запись."""
    table = df.copy() if df is not None else read_file(source, filename=filename)
    if table.empty:
        raise ValueError("Файл не содержит строк.")

    city_col = column or detect_city_column(table.columns)
    if not city_col or city_col not in table.columns:
        raise ValueError(
            "Не удалось определить столбец с городом. "
            "Укажите его явно через --column / выбор в интерфейсе."
        )

    user_map = load_user_localities()
    originals = table[city_col].tolist()
    matches = [match_city(value, threshold=threshold, user_map=user_map) for value in originals]

    result_df = table.copy()
    insert_at = list(result_df.columns).index(city_col) + 1
    extra_cols = [STATUS_COL, VARIANTS_COL]
    if add_original:
        extra_cols.append(f"{city_col}_оригинал")
    for name in extra_cols:
        if name in result_df.columns and name != city_col:
            result_df = result_df.drop(columns=[name])

    result_df[city_col] = [m.value for m in matches]
    result_df.insert(insert_at, STATUS_COL, [m.status for m in matches])
    result_df.insert(insert_at + 1, VARIANTS_COL, [m.variants_text for m in matches])
    if add_original:
        result_df.insert(insert_at + 2, f"{city_col}_оригинал", originals)

    result_df = _sort_by_confidence(result_df)

    statuses = [m.status for m in matches]
    certain = sum(s == STATUS_CERTAIN for s in statuses)
    uncertain = sum(s == STATUS_UNCERTAIN for s in statuses)
    unmatched = sum(s == STATUS_UNMATCHED for s in statuses)
    changed = sum(
        (not _is_empty(orig) and str(match.value).strip() != str(orig).strip())
        for orig, match in zip(originals, matches)
    )
    exact = sum(
        m.method in {"exact", "official", "user"} and m.status == STATUS_CERTAIN
        for m in matches
    )
    fuzzy = sum(m.method == "fuzzy" for m in matches)

    unmatched_rows = result_df.loc[result_df[STATUS_COL] == STATUS_UNMATCHED].copy()
    uncertain_rows = result_df.loc[result_df[STATUS_COL] == STATUS_UNCERTAIN].copy()

    written: Path | None = None
    if not dry_run and output is not None:
        written = write_file(result_df, output)

    return ProcessResult(
        df=result_df,
        column=city_col,
        total=len(result_df),
        changed=changed,
        unmatched=unmatched,
        uncertain=uncertain,
        certain=certain,
        exact=exact,
        fuzzy=fuzzy,
        unmatched_rows=unmatched_rows,
        uncertain_rows=uncertain_rows,
        output_path=written,
        matches=matches,
        originals=originals,
    )


def unique_unresolved_from_matches(
    originals: Iterable[object],
    matches: list[MatchResult],
) -> pd.DataFrame:
    """Уникальные НП, которые нужно уточнить вручную."""
    counts: dict[str, int] = {}
    pending: dict[str, dict[str, object]] = {}
    for original, match in zip(originals, matches):
        if match.status not in {STATUS_UNMATCHED, STATUS_UNCERTAIN}:
            continue
        source = match.cleaned or (
            "" if _is_empty(original) else strip_prefix(str(original))
        )
        key = _fold(source)
        if not key:
            continue
        counts[key] = counts.get(key, 0) + 1
        if key not in pending:
            pending[key] = {
                "населённый пункт": source,
                "статус": match.status,
                "варианты": match.variants_text,
                "район": "не выбран",
            }
    rows = []
    for key, item in pending.items():
        item["строк"] = counts[key]
        rows.append(item)
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    status_rank = frame["статус"].map({STATUS_UNCERTAIN: 0, STATUS_UNMATCHED: 1})
    return (
        frame.assign(_r=status_rank)
        .sort_values(["_r", "строк"], ascending=[True, False])
        .drop(columns="_r")
        .reset_index(drop=True)
    )


def _default_output(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_normalized{input_path.suffix}")


def _print_stats(stats: ProcessResult, output: Path | None, dry_run: bool) -> None:
    print(f"Столбец: {stats.column}")
    print(f"Всего строк: {stats.total}")
    print(f"Уверен: {stats.certain}")
    print(f"Не уверен: {stats.uncertain}")
    print(f"Не сопоставлено: {stats.unmatched}")
    print(f"Изменено: {stats.changed}")
    print(f"Точных совпадений: {stats.exact}")
    print(f"Нечётких совпадений: {stats.fuzzy}")
    if dry_run:
        print("Режим --dry-run: файл не записан.")
    elif output is not None:
        print(f"Записано: {output}")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Нормализация столбца «город» к коротким названиям "
            "муниципальных образований Краснодарского края."
        )
    )
    parser.add_argument("input", help="Входной файл: xlsx, xls, xlsm или csv")
    parser.add_argument(
        "--output",
        "-o",
        help="Путь для сохранения. По умолчанию: <имя>_normalized.<расширение>",
    )
    parser.add_argument(
        "--column",
        "-c",
        help="Имя столбца с городом (если автоопределение не сработало)",
    )
    parser.add_argument(
        "--add-original",
        action="store_true",
        help="Добавить столбец с исходным значением",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только статистика, без записи файла",
    )
    parser.add_argument(
        "--threshold",
        type=int,
        default=DEFAULT_THRESHOLD,
        help=f"Порог нечёткого совпадения 0–100 (по умолчанию {DEFAULT_THRESHOLD})",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if not 0 <= args.threshold <= 100:
        parser.error("--threshold должен быть в диапазоне 0–100")

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Файл не найден: {input_path}", file=sys.stderr)
        return 1

    output_path = (
        None if args.dry_run else Path(args.output) if args.output else _default_output(input_path)
    )

    try:
        stats = process_file(
            input_path,
            output=output_path,
            column=args.column,
            add_original=args.add_original,
            dry_run=args.dry_run,
            threshold=args.threshold,
        )
    except Exception as exc:  # noqa: BLE001 — пользовательский CLI
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    _print_stats(stats, output_path, args.dry_run)
    return 0


if __name__ == "__main__":
    sys.exit(main())
